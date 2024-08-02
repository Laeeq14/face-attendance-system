import xlsxwriter
import openpyxl
import os

def output(filename, sheet_name, row, id, name, status):
    filepath = f"{filename}.xlsx"
    if not os.path.isfile(filepath):
        workbook = xlsxwriter.Workbook(filepath)
        worksheet = workbook.add_worksheet(sheet_name)
        worksheet.write(0, 0, "Row")
        worksheet.write(0, 1, "ID")
        worksheet.write(0, 2, "Name")
        worksheet.write(0, 3, "Status")
        workbook.close()

    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook[sheet_name]
    next_row = worksheet.max_row + 1

    worksheet.append([next_row, id, name, status])

    workbook.save(filepath)
    return filepath
